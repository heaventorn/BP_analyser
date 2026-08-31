import os
import sys
import PyPDF2
from dotenv import load_dotenv
from openai import OpenAI
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import tkinter as tk

# ===== 启动密码验证（第一道门，不通过直接退出）=====
def _check_password():
    root = tk.Tk()
    root.title("身份验证")
    root.geometry("320x150")
    root.resizable(False, False)
    root.eval("tk::PlaceWindow . center")

    tk.Label(root, text="请输入访问密码", font=("微软雅黑", 11)).pack(pady=(22, 6))
    entry = tk.Entry(root, show="*", font=("微软雅黑", 13), width=18, justify="center")
    entry.pack(pady=4)
    entry.focus_set()

    state = {"ok": False, "pwd": ""}
    def on_ok():
        state["ok"] = True
        state["pwd"] = entry.get()
        root.destroy()
    def on_cancel():
        root.destroy()

    bf = tk.Frame(root)
    bf.pack(pady=12)
    tk.Button(bf, text="确定", width=8, command=on_ok, default="active").pack(side="left", padx=10)
    tk.Button(bf, text="取消", width=8, command=on_cancel).pack(side="left", padx=10)
    entry.bind("<Return>", lambda e: on_ok())
    entry.bind("<Escape>", lambda e: on_cancel())
    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()
    return state["ok"] and state["pwd"] == "0762"

if not _check_password():
    print("密码验证失败或已取消，程序退出。")
    sys.exit(1)
# ==========================================================

def resource_path(relative_path):
    """兼容开发环境 & pyinstaller打包exe后的资源路径（保留函数，备用）"""
    if hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ============配置区（所有参数集中在这里，方便修改）============
FIELD_MAPPING = {
    "公司全称": "公司名字",
    "所属行业": "所属行业",
    "是否实体产业": "是否实体产业",
    "当前融资轮次": "规模与融资轮次",
    "核心产品&技术": "核心专利技术",
    "综合毛利率": "利润率",
    "核心客户": "核心客户",
    "风险与应对": "隐藏风险",
    "是否蓝海市场": "是否蓝海市场",
    "是否有望成为国产龙头": "是否有望成为国产龙头",
    "是否可以补全国内产业链": "是否可以补全国内产业链",
    "预期可达到市占率": "预期可达到市占率",
    "综合评分": "综合打分",
    "计划融资金额": "计划融资金额",
    "投后估值": "投后估值",
    "资金主要用途": "资金主要用途",
    "产品量产阶段": "产品量产阶段",
    "直接对标企业": "直接对标企业",
    "是否匹配基金赛道": "是否匹配基金赛道",
    "落地重庆可行性": "落地重庆可行性",
    "能否满足返投要求": "能否满足返投要求",
    "近2年营收数据": "近2年营收数据",
    "自动备注": "自动备注"
}

CORE_COLUMNS = [
    "公司名字", "所属行业", "是否实体产业", "规模与融资轮次",
    "核心专利技术", "利润率", "核心客户", "隐藏风险",
    "是否蓝海市场", "是否有望成为国产龙头", "是否可以补全国内产业链", "预期可达到市占率",
    "综合打分",
    "计划融资金额", "投后估值", "资金主要用途", "产品量产阶段", "直接对标企业",
    "是否匹配基金赛道", "落地重庆可行性", "能否满足返投要求", "近2年营收数据", "自动备注"
]

col_width_map = {
    "公司名字": 20, "所属行业": 14, "是否实体产业": 12,
    "规模与融资轮次": 18, "核心专利技术": 22, "利润率": 10,
    "核心客户": 20, "隐藏风险": 25,
    "是否蓝海市场": 14,
    "是否有望成为国产龙头": 18,
    "是否可以补全国内产业链": 20,
    "预期可达到市占率": 14,
    "综合打分": 10,
    "计划融资金额": 16, "投后估值": 14, "资金主要用途": 20,
    "产品量产阶段": 14, "直接对标企业": 20,
    "是否匹配基金赛道": 14, "落地重庆可行性": 20, "能否满足返投要求": 16, "近2年营收数据": 20, "自动备注": 35
}

SYSTEM_PROMPT = """
你是重庆安诚私募股权（重庆国资产业基金）项目初审分析师，严格按照下面规则分析商业计划书BP，**只输出标准JSON，禁止多余解释、markdown、注释**。
基金核心约束：
1. 重点投资赛道：高端制造、新材料、生物医药、新能源、数字经济、AI、算力/光模块、高端装备、化工；其余赛道大幅扣分
2. 硬性要求：实体产业、可落地重庆、满足产业返投；如果纯软件无实体直接扣30分；完全无法落地重庆直接扣30分
3. 量产分级定义：实验室样品(0分)、样机小试(10分)、小批量试产(15分)、大规模量产(20分)
输出JSON固定字段，字段不能为空，无数据填"无"：
{
"文件名称": "",
"公司全称": "",
"所属行业": "",
"是否实体产业": "是/否",
"当前融资轮次": "",
"核心产品&技术": "",
"综合毛利率": "",
"核心客户": "",
"风险与应对": "",
"是否蓝海市场":"是/否，简短理由",
"是否有望成为国产龙头":"是/否，简短理由",
"是否可以补全国内产业链":"是/否，简短理由",
"预期可达到市占率":"给出百分比区间或者无",
"综合评分": "0-100分",
"注册地": "",
"成立时间": "",
"实控人&核心团队背景": "",
"计划融资金额": "",
"投后估值": "",
"资金主要用途": "",
"产品量产阶段": "",
"直接对标企业": "",
"是否匹配基金赛道": "是/否/部分匹配",
"落地重庆可行性": "高/中/低，说明理由",
"能否满足返投要求": "能/不能/待定",
"近2年营收数据": "",
"自动备注": "用1-2句话给出投资建议，包括核心亮点、主要风险、是否推荐进入下一轮尽调"
}
打分规则（总分100分）：
1. 赛道匹配度【20分】
- 完全属于基金重点赛道15‑20分；部分匹配8‑14分；完全不匹配0‑7分
2. 市场发展前景【20分】
- 属于蓝海赛道、有国产龙头潜力、能够补全国内产业链短板、具备可观潜在市占率，14‑20分
- 部分满足上述条件 7‑13分；市场竞争激烈、无国产替代价值 0‑6分
3. 重庆落地+产业返投可行性【20分】
- 落地可行性高、可以满足返投，本地产业协同强 14‑20分
- 落地难度中等、返投待定 7‑13分；很难落地、无法完成返投 0‑6分
4. 产品成熟度与量产水平【20分】
- 大规模量产16‑20分；小批量试产10‑15分；样机小试4‑9分；实验室样品0‑3分
- 具备核心专利技术壁垒可以额外加分；技术完全依赖外购降分
5. 财务与商业化订单能力【15分】
- 毛利率健康、已有稳定核心客户与订单 11‑15分
- 有少量客户、订单尚在拓展 5‑10分；无客户无订单 0‑4分
6. 核心团队产业背景【5分】
- 团队具备深厚产业从业、研发、创业背景3‑5分；团队偏科研无产业经验0‑2分
特殊扣分项：
①纯软件无实体产业直接扣30分
②项目完全无法落地重庆直接扣30分
分数最低扣到0分，不会出现负分。
"""
# ==========================================================

# =====修复打包读取同目录.env关键代码=====
if getattr(sys, 'frozen', False):
    # exe打包模式：取exe所在文件夹
    base_dir = os.path.dirname(sys.executable)
else:
    # 源码运行模式：取脚本所在文件夹
    base_dir = os.path.dirname(os.path.abspath(__file__))

env_file_path = os.path.join(base_dir, ".env")
load_dotenv(dotenv_path=env_file_path)

# 读取配置
api_key = os.getenv("DEEPSEEK_API_KEY")
base_url = os.getenv("DEEPSEEK_BASE_URL")
model_name = os.getenv("DEEPSEEK_MODEL")

# 环境变量校验
if not all([api_key, base_url, model_name]):
    print("错误：.env配置文件缺失或者DEEPSEEK_API_KEY / DEEPSEEK_BASE_URL / DEEPSEEK_MODEL 配置不全！")
    input("按回车退出...")
    sys.exit(1)

client = OpenAI(
    api_key=api_key,
    base_url=base_url
)
MODEL_NAME = model_name

def extract_pdf_text(pdf_path: str) -> str:
    full_text = ""
    try:
        with open(pdf_path, "rb") as f:
            pdf_reader = PyPDF2.PdfReader(f)
            for page in pdf_reader.pages:
                page_content = page.extract_text()
                if page_content:
                    full_text += page_content + "\n\n"
        return full_text
    except Exception as e:
        print(f"读取失败 {pdf_path}：{str(e)}")
        return ""

def clean_value(v):
    if isinstance(v, list):
        return "；".join([str(x) for x in v if x])
    if isinstance(v, dict):
        return "；".join([f"{k}:{val}" for k, val in v.items() if val])
    return str(v) if v is not None else ""

def _parse_llm_json(text):
    """健壮解析 LLM 返回的 JSON。逐级尝试：
    原样 -> 剥离 markdown 围栏 -> 提取 {..} 主体；每级再尝试清理尾逗号。全部失败返回 None。"""
    if not text:
        return None
    candidates = []
    candidates.append(text.strip())
    m = re.search(r"```[a-zA-Z]*\s*(.*?)```", text, re.S)
    if m:
        candidates.append(m.group(1).strip())
    t = text.strip()
    start = t.find("{")
    end = t.rfind("}")
    if start != -1 and end > start:
        candidates.append(t[start:end + 1])
    for c in candidates:
        try:
            return json.loads(c)
        except json.JSONDecodeError:
            pass
        try:
            return json.loads(re.sub(r",\s*([}\]])", r"\1", c))
        except json.JSONDecodeError:
            pass
    return None


def analyze_bp(pdf_text, pdf_name):
    resp = client.chat.completions.create(
        model=os.getenv("DEEPSEEK_MODEL"),
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": pdf_text[:12000]}
        ],
        temperature=0.1
    )
    json_str = resp.choices[0].message.content
    parsed = _parse_llm_json(json_str)
    if parsed is not None:
        return parsed
    print(f"\n====JSON解析失败====")
    print(f"原始返回：{json_str}")
    print(f"错误信息：未能从模型返回中解析出 JSON")
    return {
        "文件名称": pdf_name,
        "公司全称": "解析失败",
        "初审结论": "AI返回JSON格式错乱，需要人工复核这份BP"
    }

def generate_excel(all_results, output_dir):
    """把解析结果生成美化Excel，返回保存路径"""
    ordered_columns = CORE_COLUMNS
    cleaned_results = []
    for item in all_results:
        mapped_item = {}
        for src_key, dst_key in FIELD_MAPPING.items():
            mapped_item[dst_key] = item.get(src_key, "")
        cleaned = {col: clean_value(mapped_item.get(col, "")) for col in ordered_columns}
        cleaned_results.append(cleaned)

    # ============核心新增：按综合打分从高到低排序============
    def get_score(item):
        """提取分数数字，解析失败返回-1，自动排到表格末尾"""
        score_str = item.get("综合打分", "0分")
        try:
            # 去掉"分"字，转为整数用于排序
            score_num = int(score_str.replace("分", "").strip())
            return score_num
        except:
            # 非数字、解析失败的内容，统一排到最后
            return -1
    # 降序排序：分数高的在前，低的在后
    cleaned_results.sort(key=get_score, reverse=True)
    # ====================================================

    wb = Workbook()
    ws = wb.active
    ws.title = "BP初审汇总"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for col_idx, col_name in enumerate(ordered_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row_idx, item in enumerate(cleaned_results, 2):
        for col_idx, col_name in enumerate(ordered_columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=item[col_name])
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx, col_name in enumerate(ordered_columns, 1):
        width = col_width_map.get(col_name, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(cleaned_results) + 2):
        ws.row_dimensions[row_idx].height = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ✅使用时间戳，每次生成全新文件名，不会覆盖旧报告
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_path = os.path.join(output_dir, f"BP初审汇总报告_{now_str}.xlsx")
    wb.save(save_path)
    return save_path

def batch_run():
    input_dir = "BP"
    output_dir = "output"
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    pdf_list = [f for f in os.listdir(input_dir) if f.lower().endswith(".pdf")]
    all_results = []
    if not pdf_list:
        print(f"请把BP PDF放入 {input_dir} 文件夹！")
        return
    for pdf_name in pdf_list:
        try:
            print(f"正在分析：{pdf_name}")
            full_path = os.path.join(input_dir, pdf_name)
            text = extract_pdf_text(full_path)
            if not text:
                print(f"{pdf_name} 读取文本为空，跳过")
                continue
            res_data = analyze_bp(text, pdf_name)
            all_results.append(res_data)
            print(f"{pdf_name} 分析完成")
        except Exception as e:
            print(f"{pdf_name} 处理异常跳过，错误：{str(e)}")
            error_item = {
                "文件名称": pdf_name,
                "公司全称": "处理失败",
                "初审结论": f"程序异常：{str(e)}"
            }
            all_results.append(error_item)
            continue
    print("=====开始生成美化Excel=====")
    save_file = generate_excel(all_results, output_dir)
    print(f"\n全部分析完成，报告输出至：{save_file}")

if __name__ == "__main__":
    batch_run()
